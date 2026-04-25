from __future__ import annotations

import contextlib
import csv
import importlib.util
import io
import json
import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl


APP_ROOT = Path(__file__).resolve().parents[1]
CLI_ROOT = APP_ROOT / "cli"
if str(CLI_ROOT) not in sys.path:
    sys.path.insert(0, str(CLI_ROOT))

MODULE_PATH = CLI_ROOT / "reviewer_card_eval.py"


def load_module(name: str, path: Path):
    spec = importlib.util.spec_from_file_location(name, path)
    module = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    sys.modules[name] = module
    spec.loader.exec_module(module)
    return module


REVIEWER_CARD_EVAL = load_module("contract_review_reviewer_card_eval_cli", MODULE_PATH)
REVIEWER_CARD_EVAL_CORE = sys.modules["reviewer_card_eval_core"]


def write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_ground_truth(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["batch_id", "case_key", "expected_decision", "human_reason", "notes"],
        )
        writer.writeheader()
        writer.writerows(rows)


class ReviewerCardEvalTests(unittest.TestCase):
    def _write_pdf(self, path: Path, *, page_count: int) -> None:
        document = REVIEWER_CARD_EVAL_CORE.fitz.open()
        for page_index in range(page_count):
            page = document.new_page()
            page.insert_text((72, 72), f"page {page_index + 1}")
        path.parent.mkdir(parents=True, exist_ok=True)
        document.save(path)
        document.close()

    def test_cli_generates_template_and_scores_human_labels(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            summary_path = root / "runtime" / "batches" / "test-a" / "reviewer-card-summary.json"
            output_dir = root / "eval"
            ground_truth_path = root / "reviewer-card-ground-truth.csv"
            write_json(
                summary_path,
                {
                    "batch_id": "test-a",
                    "job_count": 3,
                    "decision_breakdown": {
                        "auto_pass": 1,
                        "review_recommended": 1,
                        "manual_required": 1,
                    },
                    "items": [
                        {
                            "job_id": "test-a-001",
                            "group_key": "case-auto",
                            "decision": "auto_pass",
                            "decision_label": "可自动通过",
                            "contract_amount": "8200元",
                            "pricing_amount": "8200元",
                            "difference": "0元",
                            "pending_item_names": [],
                            "primary_reason": "金额差异在可自动通过范围内（差额0元）。",
                            "next_actions": ["可低风险通过，建议保留本次金额核对记录。"],
                        },
                        {
                            "job_id": "test-a-002",
                            "group_key": "case-review",
                            "decision": "review_recommended",
                            "decision_label": "建议人工复核",
                            "contract_amount": "24013元",
                            "pricing_amount": "24326元",
                            "difference": "313元",
                            "pending_item_names": [],
                            "primary_reason": "系统报价更接近折前价，建议人工确认折扣口径。",
                            "next_actions": ["系统报价更接近折前价，请确认折扣口径。"],
                        },
                        {
                            "job_id": "test-a-003",
                            "group_key": "case-manual",
                            "decision": "manual_required",
                            "decision_label": "必须人工确认",
                            "contract_amount": "41085元",
                            "pricing_amount": "26325元",
                            "difference": "14760元",
                            "pending_item_names": ["衣柜组合"],
                            "primary_reason": "存在1个品项未入账，不能判断整单金额是否正确。",
                            "next_actions": ["优先确认未入账品项：衣柜组合。"],
                        },
                    ],
                },
            )
            write_ground_truth(
                ground_truth_path,
                [
                    {
                        "batch_id": "test-a",
                        "case_key": "case-auto",
                        "expected_decision": "manual_required",
                        "human_reason": "类目路线错",
                        "notes": "系统不该自动放行",
                    },
                    {
                        "batch_id": "test-a",
                        "case_key": "case-review",
                        "expected_decision": "review_recommended",
                        "human_reason": "折扣口径",
                        "notes": "",
                    },
                    {
                        "batch_id": "test-a",
                        "case_key": "case-manual",
                        "expected_decision": "review_recommended",
                        "human_reason": "未入账品项",
                        "notes": "审核员可以先复核，不必硬拦",
                    },
                ],
            )

            with contextlib.redirect_stdout(io.StringIO()):
                payload = REVIEWER_CARD_EVAL.run(
                    [
                        "--summary-path",
                        str(summary_path),
                        "--ground-truth-path",
                        str(ground_truth_path),
                        "--output-dir",
                        str(output_dir),
                        "--output-mode",
                        "json",
                    ]
                )

            report_json_path = output_dir / "reviewer-card-eval.json"
            report_markdown_path = output_dir / "reviewer-card-eval.md"
            calibration_path = output_dir / "reviewer-card-calibration.md"
            template_path = output_dir / "reviewer-card-ground-truth.template.csv"
            human_template_path = output_dir / "reviewer-card-human-review.template.csv"
            human_workbook_path = output_dir / "reviewer-card-human-review.xlsx"
            report_json_exists = report_json_path.exists()
            report_markdown_exists = report_markdown_path.exists()
            calibration_exists = calibration_path.exists()
            template_exists = template_path.exists()
            human_template_exists = human_template_path.exists()
            human_workbook_exists = human_workbook_path.exists()
            template_text = template_path.read_text(encoding="utf-8")
            with template_path.open("r", encoding="utf-8", newline="") as handle:
                template_rows = list(csv.DictReader(handle))
            with human_template_path.open("r", encoding="utf-8", newline="") as handle:
                human_template_rows = list(csv.DictReader(handle))
            workbook = openpyxl.load_workbook(human_workbook_path)
            worksheet = workbook.active
            validation_sheet = workbook["下拉选项"]
            workbook_rows = list(worksheet.iter_rows(min_row=2, values_only=True))
            report_markdown = report_markdown_path.read_text(encoding="utf-8")
            calibration_markdown = calibration_path.read_text(encoding="utf-8")

        self.assertTrue(report_json_exists)
        self.assertTrue(report_markdown_exists)
        self.assertTrue(calibration_exists)
        self.assertTrue(template_exists)
        self.assertTrue(human_template_exists)
        self.assertTrue(human_workbook_exists)
        self.assertEqual(payload["calibration_markdown_path"], str(calibration_path.resolve()))
        self.assertEqual(payload["human_review_template_path"], str(human_template_path.resolve()))
        self.assertEqual(payload["human_review_workbook_path"], str(human_workbook_path.resolve()))
        self.assertEqual(payload["summary"]["total_case_count"], 3)
        self.assertEqual(payload["summary"]["labeled_case_count"], 3)
        self.assertEqual(payload["summary"]["decision_match_count"], 1)
        self.assertEqual(payload["summary"]["false_release_count"], 1)
        self.assertEqual(payload["summary"]["over_escalation_count"], 1)
        self.assertEqual(payload["summary"]["human_reason_breakdown"]["折扣口径"], 1)
        self.assertEqual(payload["summary"]["actual_decision_breakdown"]["auto_pass"], 1)
        self.assertEqual(payload["summary"]["expected_decision_breakdown"]["manual_required"], 1)
        self.assertEqual(payload["items"][0]["case_key"], "case-auto")
        self.assertTrue(payload["items"][0]["false_release"])
        self.assertIn("case-auto", template_text)
        self.assertIn("actual_decision", template_text)
        self.assertIn("system_suggested_expected_decision", template_text)
        template_rows_by_case = {row["case_key"]: row for row in template_rows}
        self.assertEqual(template_rows_by_case["case-auto"]["system_suggested_expected_decision"], "auto_pass")
        self.assertEqual(template_rows_by_case["case-auto"]["system_suggested_human_reason"], "金额接近")
        self.assertEqual(template_rows_by_case["case-auto"]["expected_decision"], "")
        self.assertEqual(template_rows_by_case["case-auto"]["human_reason"], "")
        self.assertEqual(template_rows_by_case["case-review"]["system_suggested_expected_decision"], "review_recommended")
        self.assertEqual(template_rows_by_case["case-review"]["system_suggested_human_reason"], "折扣口径")
        self.assertEqual(template_rows_by_case["case-manual"]["system_suggested_expected_decision"], "manual_required")
        self.assertEqual(template_rows_by_case["case-manual"]["system_suggested_human_reason"], "未入账品项")
        human_rows_by_case = {row["合同标识"]: row for row in human_template_rows}
        workbook_rows_by_case = {str(row[0] or ""): row for row in workbook_rows}
        self.assertEqual(human_rows_by_case["case-auto"]["系统结论"], "可自动通过")
        self.assertEqual(human_rows_by_case["case-auto"]["审核员结论"], "")
        self.assertEqual(human_rows_by_case["case-auto"]["审核员原因"], "")
        self.assertEqual(human_rows_by_case["case-manual"]["待确认品项"], "衣柜组合")
        self.assertEqual(worksheet.title, "审核员校准")
        self.assertEqual(worksheet["A1"].value, "合同标识")
        self.assertEqual(validation_sheet.sheet_state, "hidden")
        self.assertEqual(validation_sheet["A2"].value, "可自动通过")
        self.assertEqual(validation_sheet["A4"].value, "必须人工确认")
        self.assertEqual(validation_sheet["B2"].value, "金额接近")
        self.assertEqual(validation_sheet["B11"].value, "证据不足")
        self.assertEqual(validation_sheet["B12"].value, "其他")
        self.assertEqual(len(worksheet.data_validations.dataValidation), 2)
        formulas = {validation.formula1 for validation in worksheet.data_validations.dataValidation}
        self.assertIn("='下拉选项'!$A$2:$A$4", formulas)
        self.assertIn("='下拉选项'!$B$2:$B$12", formulas)
        self.assertEqual((workbook_rows_by_case["case-auto"][7] or ""), "")
        self.assertEqual(workbook_rows_by_case["case-auto"][10], "暂无截图，先参考系统判断原因与原始合同。")
        self.assertIn("审核员决策卡验收报告", report_markdown)
        self.assertIn("误放行", report_markdown)
        self.assertIn("审核员校准说明", calibration_markdown)
        self.assertIn("打开 `reviewer-card-human-review.xlsx`", calibration_markdown)
        self.assertIn("只需要重点检查 审核员结论、审核员原因、审核备注，以及右侧证据截图", calibration_markdown)
        self.assertIn("`审核员结论` 和 `审核员原因` 可以直接用下拉选项选择", calibration_markdown)
        self.assertIn("case-auto", calibration_markdown)

    def test_cli_prefills_ground_truth_template_from_system_suggestions_when_requested(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            summary_path = root / "runtime" / "batches" / "test-prefill" / "reviewer-card-summary.json"
            output_dir = root / "eval"
            write_json(
                summary_path,
                {
                    "batch_id": "test-prefill",
                    "job_count": 2,
                    "items": [
                        {
                            "job_id": "test-prefill-001",
                            "group_key": "case-review",
                            "decision": "review_recommended",
                            "contract_amount": "24013元",
                            "pricing_amount": "24326元",
                            "difference": "313元",
                            "pending_item_names": [],
                            "primary_reason": "系统报价更接近折前价，建议人工确认折扣口径。",
                            "next_actions": ["系统报价更接近折前价，请确认折扣口径。"],
                        },
                        {
                            "job_id": "test-prefill-002",
                            "group_key": "case-manual",
                            "decision": "manual_required",
                            "contract_amount": "41085元",
                            "pricing_amount": "26325元",
                            "difference": "14760元",
                            "pending_item_names": ["衣柜组合"],
                            "primary_reason": "存在1个品项未入账，不能判断整单金额是否正确。",
                            "next_actions": ["优先确认未入账品项：衣柜组合。"],
                        },
                    ],
                },
            )

            with contextlib.redirect_stdout(io.StringIO()):
                REVIEWER_CARD_EVAL.run(
                    [
                        "--summary-path",
                        str(summary_path),
                        "--output-dir",
                        str(output_dir),
                        "--prefill-suggestions",
                    ]
                )

            template_path = output_dir / "reviewer-card-ground-truth.template.csv"
            human_template_path = output_dir / "reviewer-card-human-review.template.csv"
            human_workbook_path = output_dir / "reviewer-card-human-review.xlsx"
            with template_path.open("r", encoding="utf-8", newline="") as handle:
                template_rows = list(csv.DictReader(handle))
            with human_template_path.open("r", encoding="utf-8", newline="") as handle:
                human_template_rows = list(csv.DictReader(handle))
            workbook = openpyxl.load_workbook(human_workbook_path)
            worksheet = workbook.active
            workbook_rows = list(worksheet.iter_rows(min_row=2, values_only=True))

        template_rows_by_case = {row["case_key"]: row for row in template_rows}
        human_rows_by_case = {row["合同标识"]: row for row in human_template_rows}
        workbook_rows_by_case = {str(row[0] or ""): row for row in workbook_rows}
        self.assertEqual(template_rows_by_case["case-review"]["system_suggested_expected_decision"], "review_recommended")
        self.assertEqual(template_rows_by_case["case-review"]["expected_decision"], "review_recommended")
        self.assertEqual(template_rows_by_case["case-review"]["system_suggested_human_reason"], "折扣口径")
        self.assertEqual(template_rows_by_case["case-review"]["human_reason"], "折扣口径")
        self.assertEqual(template_rows_by_case["case-manual"]["system_suggested_expected_decision"], "manual_required")
        self.assertEqual(template_rows_by_case["case-manual"]["expected_decision"], "manual_required")
        self.assertEqual(template_rows_by_case["case-manual"]["system_suggested_human_reason"], "未入账品项")
        self.assertEqual(template_rows_by_case["case-manual"]["human_reason"], "未入账品项")
        self.assertEqual(human_rows_by_case["case-review"]["审核员结论"], "建议人工复核")
        self.assertEqual(human_rows_by_case["case-review"]["审核员原因"], "折扣口径")
        self.assertEqual(human_rows_by_case["case-manual"]["审核员结论"], "必须人工确认")
        self.assertEqual(human_rows_by_case["case-manual"]["审核员原因"], "未入账品项")
        self.assertEqual(workbook_rows_by_case["case-review"][7], "建议人工复核")
        self.assertEqual(workbook_rows_by_case["case-review"][8], "折扣口径")
        self.assertEqual(workbook_rows_by_case["case-manual"][7], "必须人工确认")
        self.assertEqual(workbook_rows_by_case["case-manual"][8], "未入账品项")

    def test_workbook_uses_attachment_amount_page_when_detail_page_is_missing(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            contract_path = root / "contracts" / "case.pdf"
            self._write_pdf(contract_path, page_count=13)
            review_path = root / "jobs" / "job-001" / "output" / "review.md"
            write_json(
                review_path.with_name("review.json"),
                {
                    "assets": [
                        {
                            "role_hint": "primary_contract",
                            "source_path": str(contract_path),
                            "extension": ".pdf",
                            "text_preview": (
                                "第1页 客户合同 1.1甲方委托乙方定制家具，合同总 金额为人民币32700元。"
                                "第2页 客户合同 "
                                "第13页 附件： 产品名称 产品编号 材质 数量 费用合计（元） "
                                "卡座书柜 202603112007001 北美樱桃木 1 32700 "
                                "合计 32700 折扣 100折 折扣后合计 32700"
                            ),
                        }
                    ],
                    "product_split": {"items": []},
                },
            )
            summary_path = root / "runtime" / "batches" / "test-evidence" / "reviewer-card-summary.json"
            output_dir = root / "eval"
            write_json(
                summary_path,
                {
                    "batch_id": "test-evidence",
                    "items": [
                        {
                            "job_id": "job-001",
                            "group_key": "case-evidence",
                            "decision": "manual_required",
                            "contract_amount": "32700元",
                            "pricing_amount": "",
                            "difference": "",
                            "pending_item_names": [],
                            "primary_reason": "未形成报价，需要人工确认。",
                            "review_path": str(review_path),
                        }
                    ],
                },
            )

            with contextlib.redirect_stdout(io.StringIO()):
                REVIEWER_CARD_EVAL.run(
                    [
                        "--summary-path",
                        str(summary_path),
                        "--output-dir",
                        str(output_dir),
                    ]
                )

            workbook = openpyxl.load_workbook(output_dir / "reviewer-card-human-review.xlsx")
            worksheet = workbook.active
            evidence_note = worksheet["K2"].value
            page_label = worksheet["M2"].value
            image_count = len(getattr(worksheet, "_images", []))

        self.assertEqual(page_label, "第13页")
        self.assertEqual(evidence_note, "已附合同第13页截图，请优先核对该页。")
        self.assertEqual(image_count, 1)

    def test_attachment_page_selector_handles_docx_text_without_page_markers(self) -> None:
        page_no = REVIEWER_CARD_EVAL_CORE._select_amount_or_attachment_page_from_text(
            "产品名称 产品编号 材质 数量 费用合计（元） 其他衣柜 20260399009001 北美樱桃木 "
            "合计 27226 折扣 98折 折扣后合计 26600"
        )

        self.assertEqual(page_no, 1)

    def test_text_output_prints_human_review_workbook_path(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            summary_path = root / "runtime" / "batches" / "test-output" / "reviewer-card-summary.json"
            output_dir = root / "eval"
            write_json(
                summary_path,
                {
                    "batch_id": "test-output",
                    "items": [
                        {
                            "job_id": "test-output-001",
                            "group_key": "case-output",
                            "decision": "auto_pass",
                            "contract_amount": "8200元",
                            "pricing_amount": "8200元",
                            "difference": "0元",
                            "pending_item_names": [],
                            "primary_reason": "金额差异在可自动通过范围内（差额0元）。",
                        }
                    ],
                },
            )

            stdout = io.StringIO()
            with contextlib.redirect_stdout(stdout):
                REVIEWER_CARD_EVAL.run(
                    [
                        "--summary-path",
                        str(summary_path),
                        "--output-dir",
                        str(output_dir),
                        "--output-mode",
                        "text",
                    ]
                )

        output = stdout.getvalue()
        self.assertIn("审核表：", output)
        self.assertIn("reviewer-card-human-review.xlsx", output)


if __name__ == "__main__":
    unittest.main()
