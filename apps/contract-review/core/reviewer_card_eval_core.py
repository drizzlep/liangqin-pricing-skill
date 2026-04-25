from __future__ import annotations

import csv
import re
import shutil
import subprocess
from pathlib import Path
from typing import Any

import fitz
from openpyxl import Workbook
from openpyxl.drawing.image import Image as WorkbookImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from batch_runtime import read_json, write_json, write_markdown


DECISION_ORDER = {
    "auto_pass": 0,
    "review_recommended": 1,
    "manual_required": 2,
}
DECISION_LABELS = {
    "auto_pass": "可自动通过",
    "review_recommended": "建议人工复核",
    "manual_required": "必须人工确认",
}
TEMPLATE_FIELDNAMES = [
    "batch_id",
    "case_key",
    "job_id",
    "actual_decision",
    "system_suggested_expected_decision",
    "system_suggested_human_reason",
    "expected_decision",
    "human_reason",
    "contract_amount",
    "pricing_amount",
    "difference",
    "pending_item_names",
    "system_reason",
    "notes",
]
HUMAN_TEMPLATE_FIELDNAMES = [
    "合同标识",
    "系统结论",
    "合同金额",
    "系统报价",
    "差额",
    "待确认品项",
    "系统判断原因",
    "审核员结论",
    "审核员原因",
    "审核备注",
]
HUMAN_WORKBOOK_HEADERS = HUMAN_TEMPLATE_FIELDNAMES + [
    "证据说明",
    "原文件路径",
    "证据页码",
    "证据截图",
]
HUMAN_REVIEW_REASON_OPTIONS = [
    "金额接近",
    "折扣口径",
    "未入账品项",
    "尺寸识别错",
    "类目路线错",
    "数量识别错",
    "材质识别错",
    "附件页理解错",
    "OCR/图片看不清",
    "证据不足",
    "其他",
]
DECISION_HUMAN_LABELS = {
    "auto_pass": "可自动通过",
    "review_recommended": "建议人工复核",
    "manual_required": "必须人工确认",
}
HUMAN_LABEL_TO_DECISION = {label: key for key, label in DECISION_HUMAN_LABELS.items()}


def build_reviewer_card_eval(
    *,
    summary_paths: list[Path],
    output_dir: Path,
    ground_truth_path: Path | None = None,
    prefill_suggestions: bool = False,
) -> dict[str, Any]:
    cases = _load_summary_cases(summary_paths)
    labels = _load_ground_truth_labels(ground_truth_path) if ground_truth_path else {}
    items = [_evaluate_case(case, labels.get(_case_key(case))) for case in cases]

    summary = _build_summary(items)
    payload = {
        "summary": summary,
        "items": sorted(items, key=lambda item: (item["batch_id"], item["case_key"])),
        "summary_paths": [str(path.expanduser().resolve()) for path in summary_paths],
        "ground_truth_path": str(ground_truth_path.expanduser().resolve()) if ground_truth_path else "",
    }

    resolved_output_dir = output_dir.expanduser().resolve()
    write_json(resolved_output_dir / "reviewer-card-eval.json", payload)
    write_markdown(resolved_output_dir / "reviewer-card-eval.md", _render_markdown(payload))
    write_markdown(resolved_output_dir / "reviewer-card-calibration.md", _render_calibration_markdown(payload))
    _write_ground_truth_template(
        resolved_output_dir / "reviewer-card-ground-truth.template.csv",
        cases,
        prefill_suggestions=prefill_suggestions,
    )
    _write_human_review_template(
        resolved_output_dir / "reviewer-card-human-review.template.csv",
        cases,
        prefill_suggestions=prefill_suggestions,
    )
    _write_human_review_workbook(
        resolved_output_dir / "reviewer-card-human-review.xlsx",
        cases,
        prefill_suggestions=prefill_suggestions,
    )
    payload["report_json_path"] = str(resolved_output_dir / "reviewer-card-eval.json")
    payload["report_markdown_path"] = str(resolved_output_dir / "reviewer-card-eval.md")
    payload["calibration_markdown_path"] = str(resolved_output_dir / "reviewer-card-calibration.md")
    payload["ground_truth_template_path"] = str(resolved_output_dir / "reviewer-card-ground-truth.template.csv")
    payload["human_review_template_path"] = str(resolved_output_dir / "reviewer-card-human-review.template.csv")
    payload["human_review_workbook_path"] = str(resolved_output_dir / "reviewer-card-human-review.xlsx")
    return payload


def _load_summary_cases(paths: list[Path]) -> list[dict[str, Any]]:
    cases: list[dict[str, Any]] = []
    for path in paths:
        payload = read_json(path.expanduser().resolve())
        batch_id = str(payload.get("batch_id") or path.parent.name).strip()
        for item in list(payload.get("items") or []):
            if not isinstance(item, dict):
                continue
            case = dict(item)
            case["batch_id"] = batch_id
            cases.append(case)
    return cases


def _load_ground_truth_labels(path: Path) -> dict[tuple[str, str], dict[str, str]]:
    labels: dict[tuple[str, str], dict[str, str]] = {}
    with path.expanduser().resolve().open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            batch_id = str(row.get("batch_id") or "").strip()
            case_key = str(row.get("case_key") or "").strip()
            if not batch_id or not case_key:
                continue
            labels[(batch_id, case_key)] = {
                "expected_decision": str(row.get("expected_decision") or "").strip(),
                "human_reason": str(row.get("human_reason") or "").strip(),
                "notes": str(row.get("notes") or "").strip(),
            }
    return labels


def _evaluate_case(case: dict[str, Any], label: dict[str, str] | None) -> dict[str, Any]:
    actual_decision = _normalize_decision(case.get("decision"))
    expected_decision = _normalize_decision((label or {}).get("expected_decision"))
    labeled = bool(expected_decision)
    false_release = bool(
        labeled
        and actual_decision == "auto_pass"
        and DECISION_ORDER[expected_decision] > DECISION_ORDER[actual_decision]
    )
    over_escalation = bool(
        labeled
        and actual_decision == "manual_required"
        and DECISION_ORDER[expected_decision] < DECISION_ORDER[actual_decision]
    )
    decision_match = bool(labeled and actual_decision == expected_decision)

    return {
        "batch_id": str(case.get("batch_id") or "").strip(),
        "case_key": str(case.get("group_key") or "").strip(),
        "job_id": str(case.get("job_id") or "").strip(),
        "actual_decision": actual_decision,
        "actual_decision_label": DECISION_LABELS.get(actual_decision, actual_decision),
        "expected_decision": expected_decision,
        "expected_decision_label": DECISION_LABELS.get(expected_decision, expected_decision) if expected_decision else "",
        "human_reason": str((label or {}).get("human_reason") or "").strip(),
        "notes": str((label or {}).get("notes") or "").strip(),
        "labeled": labeled,
        "decision_match": decision_match,
        "false_release": false_release,
        "over_escalation": over_escalation,
        "contract_amount": str(case.get("contract_amount") or "").strip(),
        "pricing_amount": str(case.get("pricing_amount") or "").strip(),
        "difference": str(case.get("difference") or "").strip(),
        "pending_item_names": [
            str(name).strip()
            for name in list(case.get("pending_item_names") or [])
            if str(name).strip()
        ],
        "system_reason": str(case.get("primary_reason") or "").strip(),
        "next_actions": [
            str(action).strip()
            for action in list(case.get("next_actions") or [])
            if str(action).strip()
        ],
    }


def _build_summary(items: list[dict[str, Any]]) -> dict[str, Any]:
    labeled_items = [item for item in items if item["labeled"]]
    decision_match_count = len([item for item in labeled_items if item["decision_match"]])
    false_release_items = [item for item in labeled_items if item["false_release"]]
    over_escalation_items = [item for item in labeled_items if item["over_escalation"]]
    return {
        "total_case_count": len(items),
        "labeled_case_count": len(labeled_items),
        "unlabeled_case_count": len(items) - len(labeled_items),
        "decision_match_count": decision_match_count,
        "decision_match_rate": _safe_rate(decision_match_count, len(labeled_items)),
        "false_release_count": len(false_release_items),
        "over_escalation_count": len(over_escalation_items),
        "actual_decision_breakdown": _count_by(items, "actual_decision"),
        "expected_decision_breakdown": _count_by(labeled_items, "expected_decision"),
        "human_reason_breakdown": _count_by(labeled_items, "human_reason"),
        "false_release_case_keys": [item["case_key"] for item in false_release_items],
        "over_escalation_case_keys": [item["case_key"] for item in over_escalation_items],
    }


def _write_ground_truth_template(path: Path, cases: list[dict[str, Any]], *, prefill_suggestions: bool = False) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=TEMPLATE_FIELDNAMES)
        writer.writeheader()
        for case in sorted(cases, key=lambda item: (str(item.get("batch_id") or ""), str(item.get("group_key") or ""))):
            suggested_decision, suggested_reason = _suggest_human_label(case)
            writer.writerow(
                {
                    "batch_id": str(case.get("batch_id") or "").strip(),
                    "case_key": str(case.get("group_key") or "").strip(),
                    "job_id": str(case.get("job_id") or "").strip(),
                    "actual_decision": _normalize_decision(case.get("decision")),
                    "system_suggested_expected_decision": suggested_decision,
                    "system_suggested_human_reason": suggested_reason,
                    "expected_decision": suggested_decision if prefill_suggestions else "",
                    "human_reason": suggested_reason if prefill_suggestions else "",
                    "contract_amount": str(case.get("contract_amount") or "").strip(),
                    "pricing_amount": str(case.get("pricing_amount") or "").strip(),
                    "difference": str(case.get("difference") or "").strip(),
                    "pending_item_names": "、".join(
                        str(name).strip()
                        for name in list(case.get("pending_item_names") or [])
                        if str(name).strip()
                    ),
                    "system_reason": str(case.get("primary_reason") or "").strip(),
                    "notes": "",
                }
            )


def _write_human_review_template(path: Path, cases: list[dict[str, Any]], *, prefill_suggestions: bool = False) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=HUMAN_TEMPLATE_FIELDNAMES)
        writer.writeheader()
        for case in sorted(cases, key=lambda item: (str(item.get("batch_id") or ""), str(item.get("group_key") or ""))):
            writer.writerow(_build_human_review_row(case, prefill_suggestions=prefill_suggestions))


def _write_human_review_workbook(path: Path, cases: list[dict[str, Any]], *, prefill_suggestions: bool = False) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "审核员校准"
    _build_human_review_validation_sheet(workbook)
    worksheet.append(HUMAN_WORKBOOK_HEADERS)
    _style_human_review_header(worksheet)

    screenshot_dir = path.parent / "_evidence"
    sorted_cases = sorted(cases, key=lambda item: (str(item.get("batch_id") or ""), str(item.get("group_key") or "")))
    for row_index, case in enumerate(sorted_cases, start=2):
        row = _build_human_review_row(case, prefill_suggestions=prefill_suggestions)
        evidence = _build_case_evidence(case, screenshot_dir=screenshot_dir)
        worksheet.append(
            [
                row["合同标识"],
                row["系统结论"],
                row["合同金额"],
                row["系统报价"],
                row["差额"],
                row["待确认品项"],
                row["系统判断原因"],
                row["审核员结论"],
                row["审核员原因"],
                row["审核备注"],
                evidence["evidence_note"],
                evidence["source_path"],
                evidence["page_label"],
                "",
            ]
        )
        _style_human_review_row(worksheet, row_index=row_index)
        screenshot_path = evidence.get("screenshot_path")
        if screenshot_path:
            image = WorkbookImage(str(screenshot_path))
            image.width = 240
            image.height = 180
            image.anchor = f"N{row_index}"
            worksheet.add_image(image)
            worksheet.row_dimensions[row_index].height = 140

    _configure_human_review_workbook_layout(worksheet)
    _configure_human_review_validations(workbook, worksheet)
    workbook.save(path)


def _render_markdown(payload: dict[str, Any]) -> str:
    summary = payload["summary"]
    lines = [
        "# 审核员决策卡验收报告",
        "",
        f"- 总样本数：`{summary['total_case_count']}`",
        f"- 已标注：`{summary['labeled_case_count']}`",
        f"- 结论命中：`{summary['decision_match_count']}` / `{summary['labeled_case_count']}`",
        f"- 误放行：`{summary['false_release_count']}`",
        f"- 过度拦截：`{summary['over_escalation_count']}`",
        "",
        "## 人工原因分布",
        "",
    ]
    for reason, count in summary["human_reason_breakdown"].items():
        lines.append(f"- {reason or '未标注原因'}：`{count}`")
    lines.extend(["", "## 逐单结果", ""])
    for item in payload["items"]:
        marker = ""
        if item["false_release"]:
            marker = " / 误放行"
        elif item["over_escalation"]:
            marker = " / 过度拦截"
        lines.append(
            f"- `{item['batch_id']}` / `{item['case_key']}` / "
            f"系统={item['actual_decision_label']} / "
            f"人工={item['expected_decision_label'] or '未标注'}{marker}"
        )
        if item["human_reason"]:
            lines.append(f"  人工原因：{item['human_reason']}")
        if item["pending_item_names"]:
            lines.append(f"  未入账品项：{'、'.join(item['pending_item_names'])}")
        if item["system_reason"]:
            lines.append(f"  系统原因：{item['system_reason']}")
    return "\n".join(lines).strip() + "\n"


def _render_calibration_markdown(payload: dict[str, Any]) -> str:
    summary = payload["summary"]
    lines = [
        "# 审核员校准说明",
        "",
        "## 怎么填",
        "",
        "- 打开 `reviewer-card-human-review.xlsx`。",
        "- 只需要重点检查 审核员结论、审核员原因、审核备注，以及右侧证据截图。",
        "- `审核员结论` 和 `审核员原因` 可以直接用下拉选项选择。",
        "- `审核员结论` 只用三种：可自动通过、建议人工复核、必须人工确认。",
        "- 如果系统预填正确，不用改；如果不正确，只改这三列。",
        "- `reviewer-card-ground-truth.template.csv` 是给开发和程序算分用的底表，审核员可以不看。",
        "",
        "## 结论口径",
        "",
        "- auto_pass：金额接近，审核员不需要再看。",
        "- review_recommended：金额不一定错，但建议人工快速复核。",
        "- manual_required：存在未入账、无法报价、明显差异或路线冲突，必须人工确认。",
        "",
        "## 优先校准",
        "",
    ]
    priority_items = _calibration_priority_items(payload["items"])
    if not priority_items:
        lines.append("- 暂无高优先级样本，按 CSV 顺序检查即可。")
    for item in priority_items:
        detail_parts = [
            f"系统={item['actual_decision_label']}",
            f"合同={item['contract_amount'] or '未识别'}",
            f"报价={item['pricing_amount'] or '未形成报价'}",
            f"差额={item['difference'] or '无法对比'}",
        ]
        if item["pending_item_names"]:
            detail_parts.append(f"未入账={'、'.join(item['pending_item_names'])}")
        lines.append(f"- `{item['batch_id']}` / `{item['case_key']}`：{'；'.join(detail_parts)}")
        if item["system_reason"]:
            lines.append(f"  系统原因：{item['system_reason']}")
    lines.extend(
        [
            "",
            "## 当前样本概况",
            "",
            f"- 总样本数：`{summary['total_case_count']}`",
            f"- 已标注：`{summary['labeled_case_count']}`",
            f"- 误放行：`{summary['false_release_count']}`",
            f"- 过度拦截：`{summary['over_escalation_count']}`",
        ]
    )
    return "\n".join(lines).strip() + "\n"


def _calibration_priority_items(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    def priority(item: dict[str, Any]) -> tuple[int, str, str]:
        if item["actual_decision"] == "auto_pass":
            return (0, item["batch_id"], item["case_key"])
        if item["pending_item_names"]:
            return (1, item["batch_id"], item["case_key"])
        if item["actual_decision"] == "manual_required":
            return (2, item["batch_id"], item["case_key"])
        return (3, item["batch_id"], item["case_key"])

    return sorted(items, key=priority)[:10]


def _case_key(case: dict[str, Any]) -> tuple[str, str]:
    return str(case.get("batch_id") or "").strip(), str(case.get("group_key") or "").strip()


def _normalize_decision(value: Any) -> str:
    decision = str(value or "").strip()
    return decision if decision in DECISION_ORDER else ""


def _suggest_human_label(case: dict[str, Any]) -> tuple[str, str]:
    actual_decision = _normalize_decision(case.get("decision"))
    pending_item_names = [
        str(name).strip()
        for name in list(case.get("pending_item_names") or [])
        if str(name).strip()
    ]
    system_reason = str(case.get("primary_reason") or "").strip()
    next_actions = "；".join(
        str(action).strip()
        for action in list(case.get("next_actions") or [])
        if str(action).strip()
    )
    evidence_text = f"{system_reason}；{next_actions}"

    if pending_item_names or "未入账" in evidence_text or "未形成报价" in evidence_text:
        return "manual_required", "未入账品项"
    if "折前价" in evidence_text or "折扣口径" in evidence_text or "折扣" in evidence_text:
        return "review_recommended", "折扣口径"
    if actual_decision == "auto_pass":
        return "auto_pass", "金额接近"
    if actual_decision == "review_recommended":
        return "review_recommended", "金额差异可复核"
    if actual_decision == "manual_required":
        return "manual_required", "金额差异偏大"
    return "", ""


def _count_by(items: list[dict[str, Any]], key: str) -> dict[str, int]:
    counts: dict[str, int] = {}
    for item in items:
        value = str(item.get(key) or "").strip()
        if not value:
            continue
        counts[value] = counts.get(value, 0) + 1
    return counts


def _safe_rate(numerator: int, denominator: int) -> float:
    if denominator <= 0:
        return 0.0
    return round(numerator / denominator, 4)


def _build_human_review_row(case: dict[str, Any], *, prefill_suggestions: bool) -> dict[str, str]:
    suggested_decision, suggested_reason = _suggest_human_label(case)
    return {
        "合同标识": str(case.get("group_key") or "").strip(),
        "系统结论": DECISION_HUMAN_LABELS.get(_normalize_decision(case.get("decision")), ""),
        "合同金额": str(case.get("contract_amount") or "").strip(),
        "系统报价": str(case.get("pricing_amount") or "").strip(),
        "差额": str(case.get("difference") or "").strip(),
        "待确认品项": "、".join(
            str(name).strip()
            for name in list(case.get("pending_item_names") or [])
            if str(name).strip()
        ),
        "系统判断原因": str(case.get("primary_reason") or "").strip(),
        "审核员结论": DECISION_HUMAN_LABELS.get(suggested_decision, "") if prefill_suggestions else "",
        "审核员原因": suggested_reason if prefill_suggestions else "",
        "审核备注": "",
    }


def _build_case_evidence(case: dict[str, Any], *, screenshot_dir: Path) -> dict[str, str]:
    review_payload = _load_case_review_payload(case)
    source_path = _resolve_primary_contract_path(review_payload)
    page_no = _select_case_evidence_page(review_payload, case)
    if source_path and not page_no and _is_renderable_contract_path(Path(source_path)):
        page_no = 1
    screenshot_path = ""
    evidence_note = "暂无截图，先参考系统判断原因与原始合同。"
    if source_path and page_no:
        rendered = _render_contract_screenshot(
            source_path=Path(source_path),
            page_no=page_no,
            output_dir=screenshot_dir,
            stem_hint=str(case.get("job_id") or case.get("group_key") or "case"),
        )
        if rendered:
            screenshot_path = str(rendered)
            evidence_note = f"已附合同第{page_no}页截图，请优先核对该页。"
    return {
        "evidence_note": evidence_note,
        "source_path": source_path,
        "page_label": f"第{page_no}页" if page_no else "",
        "screenshot_path": screenshot_path,
    }


def _load_case_review_payload(case: dict[str, Any]) -> dict[str, Any]:
    review_path_raw = str(case.get("review_path") or "").strip()
    if review_path_raw:
        review_path = Path(review_path_raw).expanduser()
        candidate = review_path.with_name("review.json")
        if candidate.exists():
            return read_json(candidate)
    reviewer_card_path_raw = str(case.get("reviewer_card_path") or "").strip()
    if reviewer_card_path_raw:
        reviewer_card_path = Path(reviewer_card_path_raw).expanduser()
        candidate = reviewer_card_path.with_name("review.json")
        if candidate.exists():
            return read_json(candidate)
    return {}


def _resolve_primary_contract_path(review_payload: dict[str, Any]) -> str:
    assets = list(review_payload.get("assets") or [])
    for asset in assets:
        if not isinstance(asset, dict):
            continue
        if str(asset.get("role_hint") or "").strip() == "primary_contract":
            return str(asset.get("source_path") or "").strip()
    for asset in assets:
        if not isinstance(asset, dict):
            continue
        source_path = str(asset.get("source_path") or "").strip()
        if source_path:
            return source_path
    return ""


def _select_case_evidence_page(review_payload: dict[str, Any], case: dict[str, Any]) -> int | None:
    split_items = list(review_payload.get("product_split", {}).get("items") or [])
    if not split_items:
        return _select_amount_or_attachment_page(review_payload)

    pending_names = [
        str(name).strip()
        for name in list(case.get("pending_item_names") or [])
        if str(name).strip()
    ]
    for name in pending_names:
        matched = _find_split_item_by_name(split_items, name)
        page_no = _extract_split_item_page(matched)
        if page_no:
            return page_no

    difference_sources = list(review_payload.get("reviewer_card", {}).get("difference_sources") or [])
    for source in difference_sources:
        if not isinstance(source, dict):
            continue
        for name in list(source.get("item_names") or []):
            matched = _find_split_item_by_name(split_items, str(name).strip())
            page_no = _extract_split_item_page(matched)
            if page_no:
                return page_no

    for item in split_items:
        page_no = _extract_split_item_page(item)
        if page_no:
            return page_no
    return _select_amount_or_attachment_page(review_payload)


def _select_amount_or_attachment_page(review_payload: dict[str, Any]) -> int | None:
    for asset in _ordered_contract_assets(review_payload):
        page_no = _select_amount_or_attachment_page_from_text(str(asset.get("text_preview") or ""))
        if page_no:
            return page_no
    return None


def _ordered_contract_assets(review_payload: dict[str, Any]) -> list[dict[str, Any]]:
    assets = [asset for asset in list(review_payload.get("assets") or []) if isinstance(asset, dict)]
    primary_assets = [
        asset
        for asset in assets
        if str(asset.get("role_hint") or "").strip() == "primary_contract"
    ]
    return primary_assets or assets


def _select_amount_or_attachment_page_from_text(text: str) -> int | None:
    text = text.strip()
    if not text:
        return None
    page_segments = list(_iter_page_text_segments(text))
    if not page_segments:
        return 1 if _has_amount_or_attachment_signal(text) else None

    scored_pages: list[tuple[int, int]] = []
    for page_no, segment in page_segments:
        score = _score_amount_or_attachment_segment(segment)
        if score > 0:
            scored_pages.append((score, page_no))
    if not scored_pages:
        return None
    scored_pages.sort(key=lambda item: (-item[0], item[1]))
    return scored_pages[0][1]


def _iter_page_text_segments(text: str) -> list[tuple[int, str]]:
    matches = list(re.finditer(r"第\s*(\d+)\s*页", text))
    segments: list[tuple[int, str]] = []
    for index, match in enumerate(matches):
        try:
            page_no = int(match.group(1))
        except ValueError:
            continue
        next_start = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        segments.append((page_no, text[match.start():next_start]))
    return segments


def _has_amount_or_attachment_signal(text: str) -> bool:
    return _score_amount_or_attachment_segment(text) > 0


def _score_amount_or_attachment_segment(text: str) -> int:
    normalized = re.sub(r"\s+", "", text)
    table_score = 0
    if "附件" in normalized:
        table_score += 1
    for signal in ("产品名称", "产品编号", "费用合计", "折扣后合计", "材质", "数量"):
        if signal in normalized:
            table_score += 2
    for signal in ("合计", "折扣"):
        if signal in normalized:
            table_score += 1
    if table_score >= 4:
        return 100 + table_score

    amount_score = 0
    for signal in ("合同总金额", "合同金额", "金额为人民币", "总金额"):
        if signal in normalized:
            amount_score += 1
    return 10 + amount_score if amount_score else 0


def _find_split_item_by_name(items: list[dict[str, Any]], name: str) -> dict[str, Any] | None:
    target = name.strip()
    if not target:
        return None
    for item in items:
        product_name = str(item.get("product_name") or "").strip()
        if product_name == target or target in product_name or product_name in target:
            return item
    return None


def _extract_split_item_page(item: dict[str, Any] | None) -> int | None:
    if not isinstance(item, dict):
        return None
    detail_resolution = item.get("detail_resolution") or {}
    detail_page_no = detail_resolution.get("detail_page_no")
    if isinstance(detail_page_no, int) and detail_page_no > 0:
        return detail_page_no
    boundary_start_page = item.get("boundary_start_page")
    if isinstance(boundary_start_page, int) and boundary_start_page > 0:
        return boundary_start_page
    linked_range = detail_resolution.get("linked_contract_page_range") or {}
    start_page = linked_range.get("start")
    if isinstance(start_page, int) and start_page > 0:
        return start_page
    return None


def _render_contract_screenshot(source_path: Path, *, page_no: int, output_dir: Path, stem_hint: str) -> Path | None:
    pdf_path = _resolve_renderable_pdf(source_path, output_dir=output_dir)
    if not pdf_path:
        return None
    return _render_pdf_screenshot(
        pdf_path=pdf_path,
        page_no=page_no,
        output_dir=output_dir,
        stem_hint=stem_hint,
    )


def _resolve_renderable_pdf(source_path: Path, *, output_dir: Path) -> Path | None:
    if not source_path.exists():
        return None
    suffix = source_path.suffix.lower()
    if suffix == ".pdf":
        return source_path
    if suffix in {".doc", ".docx"}:
        return _convert_document_to_pdf(source_path, output_dir=output_dir / "_converted_pdf")
    return None


def _is_renderable_contract_path(source_path: Path) -> bool:
    return source_path.suffix.lower() in {".pdf", ".doc", ".docx"}


def _convert_document_to_pdf(source_path: Path, *, output_dir: Path) -> Path | None:
    soffice = shutil.which("soffice")
    if not soffice:
        return None
    output_dir.mkdir(parents=True, exist_ok=True)
    expected_path = output_dir / f"{source_path.stem}.pdf"
    if expected_path.exists():
        return expected_path
    try:
        subprocess.run(
            [
                soffice,
                "--headless",
                "--convert-to",
                "pdf",
                "--outdir",
                str(output_dir),
                str(source_path),
            ],
            check=False,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            timeout=60,
        )
    except Exception:
        return None
    if expected_path.exists():
        return expected_path
    candidates = sorted(output_dir.glob("*.pdf"), key=lambda path: path.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else None


def _render_pdf_screenshot(pdf_path: Path, *, page_no: int, output_dir: Path, stem_hint: str) -> Path | None:
    if not pdf_path.exists() or pdf_path.suffix.lower() != ".pdf" or page_no <= 0:
        return None
    output_dir.mkdir(parents=True, exist_ok=True)
    safe_stem = re.sub(r"[^A-Za-z0-9._-]+", "-", stem_hint).strip("-") or "evidence"
    output_path = output_dir / f"{safe_stem}-p{page_no}.png"
    if output_path.exists():
        return output_path
    try:
        with fitz.open(pdf_path) as document:
            if page_no - 1 >= document.page_count:
                return None
            page = document.load_page(page_no - 1)
            pixmap = page.get_pixmap(matrix=fitz.Matrix(1.6, 1.6), alpha=False)
            pixmap.save(output_path)
    except Exception:
        return None
    return output_path


def _style_human_review_header(worksheet: Any) -> None:
    fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    font = Font(bold=True)
    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_human_review_row(worksheet: Any, *, row_index: int) -> None:
    for cell in worksheet[row_index]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def _configure_human_review_workbook_layout(worksheet: Any) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    widths = {
        "A": 36,
        "B": 14,
        "C": 12,
        "D": 12,
        "E": 10,
        "F": 20,
        "G": 42,
        "H": 16,
        "I": 16,
        "J": 20,
        "K": 26,
        "L": 48,
        "M": 10,
        "N": 34,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width


def _build_human_review_validation_sheet(workbook: Workbook) -> None:
    if "下拉选项" in workbook.sheetnames:
        options_sheet = workbook["下拉选项"]
        options_sheet.delete_rows(1, options_sheet.max_row)
    else:
        options_sheet = workbook.create_sheet("下拉选项")
    options_sheet.sheet_state = "hidden"
    options_sheet["A1"] = "审核员结论"
    options_sheet["B1"] = "审核员原因"
    for row_index, value in enumerate(DECISION_HUMAN_LABELS.values(), start=2):
        options_sheet[f"A{row_index}"] = value
    for row_index, value in enumerate(HUMAN_REVIEW_REASON_OPTIONS, start=2):
        options_sheet[f"B{row_index}"] = value


def _configure_human_review_validations(workbook: Workbook, worksheet: Any) -> None:
    options_sheet = workbook["下拉选项"]
    decision_validation = DataValidation(
        type="list",
        formula1=f"='{options_sheet.title}'!$A$2:$A${len(DECISION_HUMAN_LABELS) + 1}",
        allow_blank=True,
    )
    decision_validation.prompt = "请选择审核员结论。"
    decision_validation.error = "审核员结论请从下拉选项中选择。"
    worksheet.add_data_validation(decision_validation)
    decision_validation.add("H2:H1048576")

    reason_validation = DataValidation(
        type="list",
        formula1=f"='{options_sheet.title}'!$B$2:$B${len(HUMAN_REVIEW_REASON_OPTIONS) + 1}",
        allow_blank=True,
    )
    reason_validation.prompt = "优先从下拉里选择最接近的原因。"
    reason_validation.error = "审核员原因请优先从下拉选项中选择。"
    worksheet.add_data_validation(reason_validation)
    reason_validation.add("I2:I1048576")
